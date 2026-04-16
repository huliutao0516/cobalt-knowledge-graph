from __future__ import annotations

import csv
import json
import os
from collections import OrderedDict
from pathlib import Path
from typing import Dict, List, Optional

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parent.parent
OUTPUT_PATH = ROOT / "data" / "tenke_kfm_override.json"
DEFAULT_WORKBOOK = Path.home() / "Downloads" / "Tenke_Kisanfu_extended_trace.xlsx"

TRACE_COLUMNS = [
    ("source", 1, 2, 3),
    ("Artisanal mining", 4, 5, 6),
    ("Recycling", 7, 8, 9),
    ("Artisanal processing", 10, 11, 12),
    ("Smelting", 13, 14, 15),
    ("Trading", 16, 17, 18),
    ("Refining", 19, 20, 21),
    ("Precursor manufacturing", 22, 23, 24),
    ("Cathode manufacturing", 25, 26, 27),
    ("Battery cell manufacturing", 28, 29, 30),
    ("Battery pack manufacturing", 31, 32, 33),
    ("Electric car/scooter manufacturing", 34, 35, 36),
]

REPLACE_COMPANY_IDS = [
    "company::tenke-fungurume-mining-tfm",
    "company::kimin-sas-kisanfu-mining-sas",
    "company::kisanfu-mine-kfm",
]

WORKBOOK_SOURCE = {
    "id": "source::tenke-kisanfu-extended-trace",
    "url": "local:Tenke_Kisanfu_extended_trace.xlsx",
    "host": "Workbook import",
}

SOURCE_OVERRIDES = {
    "Tenke Fungurume": {
        "company_id": "company::tenke-fungurume-mining-tfm",
        "company_name": "Tenke Fungurume Mining (TFM)",
        "facility_id": "",
        "facility_name": "",
        "notes": [],
    },
    "Kisanfu": {
        "company_id": "company::kisanfu-mine-kfm",
        "company_name": "Kisanfu mine (KFM)",
        "facility_id": "facility::72b5ebd4ab2e",
        "facility_name": "Mine",
        "notes": ["Workbook source operator: KIMIN SAS (Kisanfu Mining SAS)"],
    },
}


def to_float(value: object) -> Optional[float]:
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    return float(text)


def split_multi_value(text: object) -> List[str]:
    if text is None:
        return []
    return [item.strip() for item in str(text).split(";") if item and item.strip()]


def infer_terminal_stage(name: str) -> str:
    lowered = name.lower()
    if "scooter" in lowered or "motorcycle" in lowered:
        return "Electric scooter manufacturing"
    return "Electric car manufacturing"


def load_company_rows() -> Dict[str, dict]:
    with (ROOT / "companies.csv").open("r", encoding="utf-8-sig", newline="") as handle:
        return {row["name"]: row for row in csv.DictReader(handle)}


def build_stage_nodes(row: tuple, company_rows: Dict[str, dict]) -> List[dict]:
    stage_nodes: List[dict] = []
    mine_label = str(row[0]).strip()
    source_override = SOURCE_OVERRIDES[mine_label]

    source_lon = to_float(row[2])
    source_lat = to_float(row[3])
    stage_nodes.append(
        {
            "stage": "Mining",
            "nodes": [
                {
                    "companyId": source_override["company_id"],
                    "companyName": source_override["company_name"],
                    "facilityId": source_override["facility_id"],
                    "facilityName": source_override["facility_name"],
                    "lat": source_lat,
                    "lon": source_lon,
                    "origin": "facility" if source_override["facility_id"] else "company",
                }
            ],
        }
    )

    for stage_name, name_idx, lon_idx, lat_idx in TRACE_COLUMNS[1:]:
        names = split_multi_value(row[name_idx])
        if not names:
            continue
        lons = [to_float(item) for item in split_multi_value(row[lon_idx])]
        lats = [to_float(item) for item in split_multi_value(row[lat_idx])]

        nodes = []
        for idx, name in enumerate(names):
            company = company_rows[name]
            node_stage = infer_terminal_stage(name) if stage_name == "Electric car/scooter manufacturing" else stage_name
            nodes.append(
                {
                    "stage": node_stage,
                    "companyId": company["node_id"],
                    "companyName": name,
                    "facilityId": "",
                    "facilityName": "",
                    "lat": lats[idx] if idx < len(lats) else to_float(company["lat"]),
                    "lon": lons[idx] if idx < len(lons) else to_float(company["lon"]),
                    "origin": "company",
                    "country": company["country_name"],
                }
            )

        stage_nodes.append({"stage": nodes[0]["stage"], "nodes": nodes})

    return stage_nodes


def build_transactions(workbook_path: Path) -> List[dict]:
    company_rows = load_company_rows()
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    sheet = workbook["trace"]

    raw_transactions: List[dict] = []

    for _, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        mine_label = str(row[0]).strip()
        source_override = SOURCE_OVERRIDES[mine_label]
        source_notes = [
            "Workbook import: Tenke_Kisanfu_extended_trace.xlsx",
            f"Workbook mine: {mine_label}",
            *source_override["notes"],
        ]

        stage_nodes = build_stage_nodes(row, company_rows)
        for stage_index in range(len(stage_nodes) - 1):
            left = stage_nodes[stage_index]
            right = stage_nodes[stage_index + 1]
            for supplier in left["nodes"]:
                for buyer in right["nodes"]:
                    raw_transactions.append(
                        {
                            "supplierCompanyId": supplier["companyId"],
                            "supplierCompany": supplier["companyName"],
                            "buyerCompanyId": buyer["companyId"],
                            "buyerCompany": buyer["companyName"],
                            "supplierFacilityId": supplier["facilityId"],
                            "supplierFacility": supplier["facilityName"],
                            "buyerFacilityId": buyer["facilityId"],
                            "buyerFacility": buyer["facilityName"],
                            "supplierCountry": company_rows.get(supplier["companyName"], {}).get("country_name", ""),
                            "buyerCountry": company_rows.get(buyer["companyName"], {}).get("country_name", ""),
                            "supplierStage": left["stage"],
                            "buyerStage": right["stage"],
                            "chainLink": f"{left['stage']} to {right['stage']}",
                            "inputCommodityIds": [],
                            "inputCommodities": [],
                            "outputCommodityIds": [],
                            "outputCommodities": [],
                            "sourceIds": [WORKBOOK_SOURCE["id"]],
                            "sourceCount": 1,
                            "amountTonnesRaw": "",
                            "amountTonnesValue": None,
                            "amountUnitsRaw": "",
                            "amountUnitsValue": None,
                            "amountUsdRaw": "",
                            "amountUsdValue": None,
                            "amountYuanRaw": "",
                            "amountYuanValue": None,
                            "amountEnergyRaw": "",
                            "amountEnergyValue": None,
                            "date": "",
                            "expectedDate": "",
                            "realised": "",
                            "notes": source_notes,
                            "hasQuantity": False,
                            "hasDate": False,
                            "sourceLat": supplier["lat"],
                            "sourceLon": supplier["lon"],
                            "sourcePointOrigin": supplier["origin"],
                            "targetLat": buyer["lat"],
                            "targetLon": buyer["lon"],
                            "targetPointOrigin": buyer["origin"],
                        }
                    )

    deduped: "OrderedDict[tuple, dict]" = OrderedDict()
    for tx in raw_transactions:
        # The workbook trace can surface Umicore in a "Smelting" column, but the
        # published TFM/KFM chain should start Umicore at refining. Keep those
        # smelting-stage Umicore branches out of the override snapshot so the
        # focused chain view does not fabricate a false Umicore smelting node.
        if tx["supplierCompanyId"] == "company::umicore-s-a" and tx["supplierStage"] == "Smelting":
            continue
        if tx["buyerCompanyId"] == "company::umicore-s-a" and tx["buyerStage"] == "Smelting":
            continue

        key = (
            tx["supplierCompanyId"],
            tx["supplierFacilityId"],
            tx["buyerCompanyId"],
            tx["buyerFacilityId"],
            tx["supplierStage"],
            tx["buyerStage"],
            tuple(tx["sourceIds"]),
            tuple(tx["notes"]),
        )
        current = deduped.get(key)
        if current is None:
            current = {**tx, "_pathCount": 0}
            deduped[key] = current
        current["_pathCount"] += 1

    transactions: List[dict] = []
    for tx_index, tx in enumerate(deduped.values(), start=1):
        path_count = tx.pop("_pathCount")
        notes = list(tx["notes"])
        if path_count > 1:
            notes.append(f"Workbook path count: {path_count}")
        tx_id = f"transaction::override-tenke-kfm-{tx_index:05d}"
        transactions.append(
            {
                "id": tx_id,
                "nodeId": tx_id,
                **tx,
                "notes": notes,
            }
        )

    return transactions


def main() -> None:
    workbook_path = Path(os.environ.get("TENKE_KISANFU_XLSX", DEFAULT_WORKBOOK))
    if not workbook_path.exists():
        raise FileNotFoundError(f"Workbook not found: {workbook_path}")

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    payload = {
        "replaceCompanyIds": REPLACE_COMPANY_IDS,
        "sources": [WORKBOOK_SOURCE],
        "entities": [],
        "operatorPairs": [
            {
                "companyId": "company::kisanfu-mine-kfm",
                "facilityId": "facility::72b5ebd4ab2e",
                "side": "supplier",
            }
        ],
        "transactions": build_transactions(workbook_path),
    }

    OUTPUT_PATH.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    print(f"Wrote {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
